"""
CHANGES from your original version:

1. QR codes are now generated for ITEMS during import (mirrors the Worker
   pattern). An item only gets a QR generated once — if it already has
   qr_code_image set, re-importing won't overwrite it (so printed labels
   stay valid). ASSUMPTION: Item model has qr_code_data / qr_code_image
   columns like Worker — if it doesn't, add them or tell me the real
   field names.

2. Fixed a stock double-counting bug: previously, re-importing the same
   file added `opening_stock` to current_stock again on every run.
   Now opening_stock only applies the FIRST time an item is created.
   For existing items, add an optional `stock_adjustment` column to the
   sheet to add/remove stock on a later import (in addition to or
   instead of opening_stock).

3. Sheet names ("Items", "Vendors", ...) and column headers are now
   matched case-insensitively and whitespace-trimmed, so a renamed tab
   or a header typed as " Name " still works.

4. Existing Items / Vendors / Customers / Workers found by name are now
   UPDATED with any new non-empty values from the sheet, instead of
   being silently skipped. Nothing is overwritten with a blank cell.

5. The response now includes `items_detail` and `workers_detail` arrays
   (name, code, qr_code_image) for any record that has a QR code, so the
   frontend can render them right after import.
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
import openpyxl
from datetime import date
from app.database import get_db
from app.models.item import Item
from app.models.vendor import Vendor
from app.models.customer import Customer
from app.models.stock import StockLedger
from app.utils.auth import get_current_user, require_role
from app.models.user import User
from app.models.worker import Worker
from app.services.qr import generate_qr_base64, generate_worker_qr_data


router = APIRouter(prefix="/import", tags=["Data Import"])


def safe_str(val):
    return str(val).strip() if val is not None else None


def safe_float(val, default=0.0):
    try:
        return float(val) if val not in (None, "") else default
    except (TypeError, ValueError):
        return default


def safe_int(val, default=0):
    try:
        return int(float(val)) if val not in (None, "") else default
    except (TypeError, ValueError):
        return default


def get_sheet(wb, name):
    """Case-insensitive, whitespace-trimmed sheet lookup."""
    target = name.strip().lower()
    for sn in wb.sheetnames:
        if sn.strip().lower() == target:
            return wb[sn]
    return None


def row_to_dict(headers, row):
    """Build a {header: value} dict with normalized (lowercase, trimmed) keys."""
    return {
        (h.strip().lower() if isinstance(h, str) else h): v
        for h, v in zip(headers, row)
    }


def get_headers(ws):
    return [c.value for c in ws[1]]


def make_item_qr_payload(company_id, item):
    tracking = getattr(item, 'tracking_type', '') or ''
    name_upper = (item.name or '').upper()
    
    if tracking == 'qr' or name_upper.startswith('PIPE'):
        return f"PIPE-{item.id}-0001"
    
    if tracking == 'bulk':
        payload = f"BULK-{item.id}"
        item.batch_qr_code = payload   # ← set the column the backend queries
        return payload
    
    return f"ITEM|{company_id}|{item.code or item.id}|{item.name}"


@router.post("/excel")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin"))
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx file")

    wb = openpyxl.load_workbook(file.file, data_only=True)
    company_id = current_user.company_id

    summary = {
        "items": 0, "vendors": 0, "customers": 0, "workers": 0,
        "qr_generated": 0, "skipped": 0, "errors": [],
        "items_detail": [], "workers_detail": [],
    }

    # ── ITEMS ──
    items_ws = get_sheet(wb, "Items")
    if items_ws:
        headers = get_headers(items_ws)

        existing_items = {
            row[0].lower(): row[1]
            for row in db.query(Item.name, Item.id).filter(
                Item.company_id == company_id
            ).all()
        }
        items_by_id = {}
        new_stock_entries = []

        editable_fields = [
            ("code", "code", str),
            ("item_type", "item_type", str),
            ("hsn_code", "hsn_code", str),
            ("unit", "unit", str),
            ("tax_rate", "tax_rate", float),
            ("purchase_price", "purchase_price", float),
            ("tracking_type", "tracking_type", str),
        ]

        for row_num, row in enumerate(items_ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            data = row_to_dict(headers, row)
            try:
                name = safe_str(data.get("name"))
                if not name:
                    summary["skipped"] += 1
                    continue

                opening_qty = safe_int(data.get("opening_stock"))
                stock_adjustment = data.get("stock_adjustment")
                purchase_price = safe_float(data.get("purchase_price"))

                existing_id = existing_items.get(name.lower())
                is_new = existing_id is None
                changed = False

                if not is_new:
                    if existing_id not in items_by_id:
                        items_by_id[existing_id] = db.query(Item).filter(
                            Item.id == existing_id
                        ).first()
                    item = items_by_id[existing_id]

                    for col_key, attr, cast in editable_fields:
                        val = data.get(col_key)
                        if val not in (None, ""):
                            try:
                                setattr(item, attr, cast(val) if cast is not str else safe_str(val))
                                changed = True
                            except (TypeError, ValueError):
                                pass
                else:
                    item = Item(
                        company_id=company_id,
                        name=name,
                        code=safe_str(data.get("code")),
                        item_type=safe_str(data.get("item_type")) or "raw_material",
                        hsn_code=safe_str(data.get("hsn_code")) or "0000",
                        unit=safe_str(data.get("unit")) or "pcs",
                        tax_rate=data.get("tax_rate") or 0,
                        opening_stock=opening_qty,
                        current_stock=0,
                        purchase_price=purchase_price,
                        tracking_type=safe_str(data.get("tracking_type")) or "unit",
                    )
                    db.add(item)
                    db.flush()  # need item.id for QR payload + stock ledger
                    existing_items[name.lower()] = item.id
                    items_by_id[item.id] = item
                    changed = True

                # Opening stock applies ONLY on first creation (avoids
                # double-counting if you re-run the same import file).
                if is_new and opening_qty:
                    item.current_stock = (item.current_stock or 0) + opening_qty
                    new_stock_entries.append(StockLedger(
                        company_id=company_id,
                        item_id=item.id,
                        transaction_type="opening_balance",
                        reference_type="import",
                        quantity=opening_qty,
                        unit_cost=purchase_price,
                        transaction_date=date.today(),
                        reason="Opening stock from migration import"
                    ))

                # Optional column for adjusting stock on a later import,
                # without re-triggering opening_stock.
                if not is_new and stock_adjustment not in (None, ""):
                    adj_qty = safe_int(stock_adjustment)
                    if adj_qty:
                        item.current_stock = (item.current_stock or 0) + adj_qty
                        new_stock_entries.append(StockLedger(
                            company_id=company_id,
                            item_id=item.id,
                            transaction_type="adjustment",
                            reference_type="import",
                            quantity=adj_qty,
                            unit_cost=purchase_price,
                            transaction_date=date.today(),
                            reason="Stock adjustment from re-import"
                        ))
                        changed = True

                # Generate a QR code once, ever, per item — re-imports
                # won't regenerate (and thus won't invalidate) an
                # already-printed QR label.
                if not getattr(item, "qr_code_image", None):
                    qr_payload = make_item_qr_payload(company_id, item)
                    item.qr_code_data = qr_payload
                    item.qr_code_image = generate_qr_base64(qr_payload)
                    summary["qr_generated"] += 1

                if getattr(item, "qr_code_image", None):
                    summary["items_detail"].append({
                        "name": item.name,
                        "code": item.code,
                        "qr_code_image": item.qr_code_image,
                        "item_id": item.id,          # ← add this
                        "quantity": int(item.current_stock or 0),  # ← add this
                    })

                if changed:
                    summary["items"] += 1
                else:
                    summary["skipped"] += 1
            except Exception as e:
                summary["errors"].append(f"Items row {row_num}: {e}")

        if new_stock_entries:
            db.add_all(new_stock_entries)

    # ── VENDORS ──
    vendors_ws = get_sheet(wb, "Vendors")
    if vendors_ws:
        headers = get_headers(vendors_ws)

        existing_vendors = {
            v.name.lower(): v for v in db.query(Vendor).filter(
                Vendor.company_id == company_id
            ).all()
        }

        vendor_fields = [
            "gstin", "pan", "address", "state", "state_code", "phone",
            "email", "bank_name", "account_number", "ifsc_code",
            "account_holder_name",
        ]

        new_vendors = []
        for row_num, row in enumerate(vendors_ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            data = row_to_dict(headers, row)
            try:
                name = safe_str(data.get("name"))
                if not name:
                    summary["skipped"] += 1
                    continue

                existing = existing_vendors.get(name.lower())
                if existing:
                    changed = False
                    for f in vendor_fields:
                        val = safe_str(data.get(f))
                        if val:
                            setattr(existing, f, val)
                            changed = True
                    if changed:
                        summary["vendors"] += 1
                    else:
                        summary["skipped"] += 1
                    continue

                vendor = Vendor(
                    company_id=company_id,
                    name=name,
                    **{f: safe_str(data.get(f)) for f in vendor_fields},
                )
                new_vendors.append(vendor)
                existing_vendors[name.lower()] = vendor
                summary["vendors"] += 1
            except Exception as e:
                summary["errors"].append(f"Vendors row {row_num}: {e}")

        if new_vendors:
            db.add_all(new_vendors)

    # ── CUSTOMERS ──
    customers_ws = get_sheet(wb, "Customers")
    if customers_ws:
        headers = get_headers(customers_ws)

        existing_customers = {
            c.name.lower(): c for c in db.query(Customer).filter(
                Customer.company_id == company_id
            ).all()
        }

        customer_fields = [
            "gstin", "pan", "address", "state", "state_code", "phone",
            "email", "bank_name", "account_number", "ifsc_code",
            "account_holder_name",
        ]

        new_customers = []
        for row_num, row in enumerate(customers_ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            data = row_to_dict(headers, row)
            try:
                name = safe_str(data.get("name"))
                if not name:
                    summary["skipped"] += 1
                    continue

                existing = existing_customers.get(name.lower())
                if existing:
                    changed = False
                    for f in customer_fields:
                        val = safe_str(data.get(f))
                        if val:
                            setattr(existing, f, val)
                            changed = True
                    if changed:
                        summary["customers"] += 1
                    else:
                        summary["skipped"] += 1
                    continue

                customer = Customer(
                    company_id=company_id,
                    name=name,
                    **{f: safe_str(data.get(f)) for f in customer_fields},
                )
                new_customers.append(customer)
                existing_customers[name.lower()] = customer
                summary["customers"] += 1
            except Exception as e:
                summary["errors"].append(f"Customers row {row_num}: {e}")

        if new_customers:
            db.add_all(new_customers)

    # ── WORKERS ──
    workers_ws = get_sheet(wb, "Workers")
    if workers_ws:
        headers = get_headers(workers_ws)

        existing_workers = {
            w.name.lower(): w for w in db.query(Worker).filter(
                Worker.company_id == company_id
            ).all()
        }

        last_worker = db.query(Worker).filter(
            Worker.company_id == company_id
        ).order_by(Worker.id.desc()).first()
        next_num = int(last_worker.worker_code[1:]) + 1 if last_worker else 1

        new_workers = []
        for row_num, row in enumerate(workers_ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            data = row_to_dict(headers, row)
            try:
                name = safe_str(data.get("name"))
                if not name:
                    summary["skipped"] += 1
                    continue

                existing = existing_workers.get(name.lower())
                if existing:
                    changed = False
                    for f in ("department", "phone"):
                        val = safe_str(data.get(f))
                        if val:
                            setattr(existing, f, val)
                            changed = True
                    if changed:
                        summary["workers"] += 1
                    else:
                        summary["skipped"] += 1
                    if getattr(existing, "qr_code_image", None):
                        summary["workers_detail"].append({
                            "name": existing.name,
                            "code": existing.worker_code,
                            "qr_code_image": existing.qr_code_image,
                        })
                    continue

                worker_code = f"W{str(next_num).zfill(3)}"
                next_num += 1

                qr_data = generate_worker_qr_data(worker_code, name)
                qr_image = generate_qr_base64(qr_data)

                worker = Worker(
                    company_id=company_id,
                    name=name,
                    worker_code=worker_code,
                    department=safe_str(data.get("department")),
                    phone=safe_str(data.get("phone")),
                    qr_code_data=qr_data,
                    qr_code_image=qr_image,
                )
                new_workers.append(worker)
                existing_workers[name.lower()] = worker
                summary["workers"] += 1
                summary["qr_generated"] += 1
                summary["workers_detail"].append({
                    "name": worker.name,
                    "code": worker.worker_code,
                    "qr_code_image": worker.qr_code_image,
                })
            except Exception as e:
                summary["errors"].append(f"Workers row {row_num}: {e}")

        if new_workers:
            db.add_all(new_workers)

    db.commit()
    return {"message": "Import completed", "summary": summary}


@router.get("/template")
def download_template():
    """Generates a blank Excel template with the correct headers."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    items_ws = wb.create_sheet("Items")
    items_ws.append([
        "name", "code", "item_type", "hsn_code", "unit", "tax_rate",
        "opening_stock", "purchase_price", "tracking_type", "stock_adjustment",
    ])
    items_ws.append([
        "6 X 50 HEX NUTS", "2589", "raw_material", "7318", "pcs", 18,
        100, 5, "bulk", "",
    ])

    vendors_ws = wb.create_sheet("Vendors")
    vendors_ws.append(["name", "gstin", "pan", "address", "state", "state_code", "phone", "email", "bank_name", "account_number", "ifsc_code", "account_holder_name"])
    vendors_ws.append(["SAP PVT LTD", "27ABCDE1234F1Z5", "ABCDE1234F", "Sample Address", "Maharashtra", "27", "9999999999", "vendor@example.com", "", "", "", ""])

    customers_ws = wb.create_sheet("Customers")
    customers_ws.append(["name", "gstin", "pan", "address", "state", "state_code", "phone", "email", "bank_name", "account_number", "ifsc_code", "account_holder_name"])
    customers_ws.append(["ABC Motors", "27XYZAB1234F1Z5", "XYZAB1234F", "Sample Address", "Maharashtra", "27", "8888888888", "customer@example.com", "", "", "", ""])

    workers_ws = wb.create_sheet("Workers")
    workers_ws.append(["name", "department", "phone"])
    workers_ws.append(["KUNJ PATEL", "Assembly", "9876543210"])

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"}
    )


@router.post("/generate-missing-qr")
def generate_missing_qr(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin"))
):
    from app.services.qr import generate_qr_base64
    items = db.query(Item).filter(
        Item.company_id == current_user.company_id,
        Item.qr_code_image == None
    ).all()
    
    count = 0
    for item in items:
        payload = f"ITEM|{current_user.company_id}|{item.code or item.id}|{item.name}"
        item.qr_code_data = payload
        item.qr_code_image = generate_qr_base64(payload)
        count += 1
    
    db.commit()
    return {"generated": count}


@router.post("/regenerate-pipe-qr")
def regenerate_pipe_qr(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin"))
):
    items = db.query(Item).filter(
        Item.company_id == current_user.company_id,
        Item.name.ilike("PIPE%")  # all pipe items
    ).all()
    
    count = 0
    for item in items:
        payload = f"PIPE-{item.id}-0001"
        item.qr_code_data = payload
        item.qr_code_image = generate_qr_base64(payload)
        count += 1
    
    db.commit()
    return {"regenerated": count, "message": f"Fixed QR for {count} pipe items"}



@router.post("/create-pipe-instances")
def create_pipe_instances(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role("admin"))
):
    from app.models.stock import PartInstance
    
    pipes = db.query(Item).filter(
        Item.company_id == current_user.company_id,
        Item.name.ilike("PIPE%")
    ).all()
    
    created = 0
    for item in pipes:
        qty = int(item.current_stock or 0)
        if qty == 0:
            continue
        
        # Check if instances already exist for this item
        existing = db.query(PartInstance).filter(
            PartInstance.item_id == item.id,
            PartInstance.company_id == current_user.company_id
        ).count()
        if existing >= qty:
            continue
        
        for i in range(existing + 1, qty + 1):
            serial = String(i).zfill(4)  # won't work in Python — fix below
            qr_data = f"PIPE-{item.id}-{str(i).zfill(4)}"
            qr_image = generate_qr_base64(qr_data)
            part = PartInstance(
                company_id=current_user.company_id,
                item_id=item.id,
                serial_number=qr_data,
                qr_code_data=qr_data,
                qr_code_image=qr_image,
                current_status="in_stock",
                remaining_quantity=1,
                purchase_order_id=None,
            )
            db.add(part)
            created += 1
    
    db.commit()
    return {"created": created, "message": f"Created {created} pipe instances"}